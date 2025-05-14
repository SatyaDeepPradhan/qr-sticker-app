from flask import Flask, render_template, request, redirect, url_for, send_file, session, flash
import qrcode
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import os
from io import BytesIO
import smtplib
from email.message import EmailMessage
import tempfile
import urllib.parse

app = Flask(__name__)
app.secret_key = 'supersecretkey'
EXCEL_FILE = 'records.xlsx'
FONT_PATH = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Owner Name", "Contact Number", "Vehicle Number", "Emergency Contact", "Relationship", "Blood Group"])
        wb.save(EXCEL_FILE)
init_excel()

def generate_sticker(data_dict):
    qr_data = " | ".join([f"{k}: {v}" for k, v in data_dict.items()])
    qr = qrcode.make(qr_data)
    sticker = Image.new("RGB", (450, 225), "black")
    qr = qr.resize((150, 150))
    sticker.paste(qr, (20, 35))
    draw = ImageDraw.Draw(sticker)
    font = ImageFont.truetype(FONT_PATH, 18)
    small_font = ImageFont.truetype(FONT_PATH, 12)
    draw.text((190, 60), f"Vehicle No:\n{data_dict['Vehicle Number']}", fill="white", font=font)
    draw.text((300, 200), "Developed by SDP", fill="white", font=small_font)
    output = BytesIO()
    sticker.save(output, format="PNG")
    output.seek(0)
    return output

def record_exists(vehicle_number):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == vehicle_number:
            return row
    return None

def add_record(data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([data[k] for k in ["Owner Name", "Contact Number", "Vehicle Number", "Emergency Contact", "Relationship", "Blood Group"]])
    wb.save(EXCEL_FILE)

def delete_record(vehicle_number):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[2].value == vehicle_number:
            ws.delete_rows(idx)
            break
    wb.save(EXCEL_FILE)

def send_email(to_email, image_data, subject="Your Vehicle QR Sticker"):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = "your_email@example.com"
    msg['To'] = to_email
    msg.set_content("Attached is your vehicle QR sticker.")
    msg.add_attachment(image_data.read(), maintype='image', subtype='png', filename='sticker.png')
    image_data.seek(0)
    with smtplib.SMTP('smtp.example.com', 587) as smtp:
        smtp.starttls()
        smtp.login("your_email@example.com", "your_password")
        smtp.send_message(msg)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        action = request.form.get('action')
        data = {k: request.form[k] for k in ["Owner Name", "Contact Number", "Vehicle Number", "Emergency Contact", "Relationship", "Blood Group"]}
        existing = record_exists(data["Vehicle Number"])
        if existing:
            flash("Record already exists. Showing existing QR code.")
        else:
            add_record(data)
            flash("QR Code Generated.")
        image = generate_sticker(data)
        if action == 'email':
            to_email = request.form.get('email')
            send_email(to_email, image)
            flash("Sticker sent via email.")
            return redirect(url_for('index'))
        elif action == 'whatsapp':
            whatsapp_url = f"https://wa.me/?text={urllib.parse.quote('Your QR sticker')}"
            flash("Sticker prepared for WhatsApp.")
            return redirect(whatsapp_url)
        else:
            return send_file(image, mimetype='image/png')
    return render_template('index.html')

@app.route('/records', methods=['GET', 'POST'])
def records():
    if 'admin' not in session:
        return redirect(url_for('login'))
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('records.html', rows=rows)

@app.route('/export')
def export():
    return send_file(EXCEL_FILE, as_attachment=True)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST' and request.form['password'] == 'admin':
        session['admin'] = True
        return redirect(url_for('records'))
    return render_template('login.html')

@app.route('/delete/<vehicle_number>')
def delete(vehicle_number):
    delete_record(vehicle_number)
    flash("Record deleted successfully.")
    return redirect(url_for('records'))

if __name__ == '__main__':
    app.run(debug=True)